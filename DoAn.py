import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import mysql.connector
from tkcalendar import DateEntry
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import warnings

warnings.filterwarnings("ignore")
# ---------------------------------------------------------------------------------------------------------------------
# Chúng em xin cảm ơn thầy đã chấm bài của chúng em! Chúc thầy sức khỏe và thành công!
# ---------------------------------------------------------------------------------------------------------------------

def mo_Doan():

    def connect_db():
        return mysql.connector.connect(
            host="127.0.0.1",
            user="root",
            password="12345",
            database="qlbaihat"        )
    def center_window(win, w=950, h=550):
        ws = win.winfo_screenwidth()
        hs = win.winfo_screenheight()
        x = (ws // 2) - (w // 2)
        y = (hs // 2) - (h // 2)
        win.geometry(f'{w}x{h}+{x}+{y}')

    def load_data():
        for i in tree.get_children():
            tree.delete(i)
        conn = connect_db()
        try:
            cur = conn.cursor()
            cur.execute("SELECT * FROM baihat")
            rows = cur.fetchall()
            for row in rows:
                tree.insert("", tk.END, values=row)
        except Exception as e:
            messagebox.showerror("Lỗi kết nối", str(e))
        finally:
            conn.close()

    def clear_input():
        entry_mabh.delete(0, tk.END)
        entry_tenbh.delete(0, tk.END)
        entry_casi.delete(0, tk.END)
        gender_var.set("Nam")
        date_entry.set_date("2000-01-01")
        cbb_loai.set("")
        entry_mabh.config(state='normal')

    def ThemBaiHat():
        mabh = entry_mabh.get()
        tenbh = entry_tenbh.get()
        gioitinh = gender_var.get()
        casi = entry_casi.get()
        ngayph = date_entry.get_date()
        loaibh = cbb_loai.get()

        if mabh == "" or tenbh == "" or casi == "":
            messagebox.showwarning("Thiếu dữ liệu", "Vui lòng nhập Mã, Họ và tên Ca Sĩ")
            return
        conn = connect_db()
        try:
            cur = conn.cursor()
            cur.execute("INSERT INTO baihat VALUES (%s, %s, %s,%s, %s, %s)",
            (mabh, tenbh, casi,gioitinh,ngayph, loaibh))
            conn.commit()
            messagebox.showinfo("Thành công", "Thêm Bài Hát thành công")
            load_data()
            clear_input()
        except mysql.connector.IntegrityError:
            messagebox.showerror("Lỗi", f"Mã Bài Hát '{mabh}' đã tồn tại!")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
        finally:
            conn.close()

    def XoaBaiHat():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Chưa chọn", "Hãy chọn Bài Hát trên bảng để xóa")
            return
        mabh = tree.item(selected)["values"][0]

        confirm = messagebox.askyesno("Xác nhận", f"Bạn có chấc muốn xóa giáo viên có mã {mabh}?")
        if confirm:
            conn = connect_db()
            cur = conn.cursor()
            cur.execute("DELETE FROM BaiHat WHERE ma_bh = %s", (mabh,))
            conn.commit()
            conn.close()
            load_data()
            clear_input()
            messagebox.showinfo("Thành công", "Đã xóa bài hát")

    def SuaBaiHat(event=None):
        selected = tree.selection()
        if not selected:
            if event is None:
                messagebox.showwarning("Chưa chọn", "Hãy chọn Bài Hát để sửa")
            return
    
        values = tree.item(selected)["values"]

        entry_mabh.delete(0, tk.END)
        entry_mabh.insert(0, values[0])
        entry_mabh.config(state='readonly')

        entry_tenbh.delete(0, tk.END)
        entry_tenbh.insert(0, values[1])

        entry_casi.delete(0, tk.END)
        entry_casi.insert(0, values[2])

        gender_var.set(values[3])
        date_entry.set_date(values[4])
        cbb_loai.set(values[5])

    def LuuBaiHat():
        mabh = entry_mabh.get()
        tenbh = entry_tenbh.get()
        casi = entry_casi.get()
        gioitinh = gender_var.get()
        ngayph = date_entry.get_date()
        loai = cbb_loai.get()

        if mabh == "":
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn Sửa một Bài Hát trước khi Lưu")
            return
        conn = connect_db()
        try:
            cur = conn.cursor()
            sql = """UPDATE baihat
                     SET tenbaihat = %s, casi = %s,gioitinh = %s,ngay_phathanh = %s, loai = %s
                     WHERE ma_bh = %s"""
            val = (tenbh,casi,gioitinh ,ngayph, loai, mabh)
            cur.execute(sql, val)
            conn.commit()
            messagebox.showinfo("Thành công", "Cập nhật thông tin Bài Hát thành công")
            load_data()
            clear_input()
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
        finally:
            conn.close()

    def TimKiem():
        search_win = tk.Toplevel(root)
        search_win.title("Tìm kiếm Bài Hát")
        search_win.geometry("300x120")

        tk.Label(search_win, text="Nhập mã hoặc tên Bài Hát:").pack(pady=10)
        entry_search = tk.Entry(search_win, width=30)
        entry_search.pack(pady=5)
        def ThucHienTim():
            keyword = entry_search.get()
            if keyword == "":
                messagebox.showwarning("Thông báo", "Vui lòng nhập từ khóa!")
                return
            for i in tree.get_children():
                tree.delete(i)
            conn = connect_db()
            try:
                cur = conn.cursor()
                sql = "SELECT * FROM baihat WHERE ma_bh LIKE %s OR tenbaihat LIKE %s OR casi LIKE %s"
                val = (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")
                cur.execute(sql, val)
                rows = cur.fetchall()
                if len(rows) == 0:
                    messagebox.showinfo("Kết quả", "Không tìm thấy bài hát nào.")
                    load_data()
                else:
                    for row in rows:
                        tree.insert("", tk.END, values=row)
                search_win.destroy()
            except Exception as e:
                messagebox.showerror("Lỗi", str(e))
            finally:
                conn.close()

        tk.Button(search_win, text="Tìm kiếm", command=ThucHienTim).pack(pady=10)

    def XuatExcel():
        conn = connect_db()
        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                     filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                     title="Lưu file Excel")
            if file_path:
                cur = conn.cursor()
                cur.execute("SELECT * FROM baihat")
                rows = cur.fetchall()
            
                wb = Workbook()
                ws = wb.active
                ws.title = "Danh Sách Bài Hát"
            
                headers = ["Mã BH", "Tên Bài Hát", "Ca Sĩ","Giới Tính",  "Ngày Phát Hành", "Loại"]
                ws.append(headers)
            
                header_font = Font(bold=True, color="FFFFFF")
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    cell.fill = openpyxl.styles.PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")

                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
            
                for row in rows:
                    ws.append(row)
                    for cell in ws[ws.max_row]:
                        cell.border = thin_border
                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = length + 5
                wb.save(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất file Excel tại:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
        finally:
            conn.close()


        
    root = tk.Tk()
    root.title("Quản lý Bài Hát")
    center_window(root)
    root.resizable(True, True)

    lbl_title = tk.Label(root, text="Quản Lý Bài Hát", font=("Arial", 20, "bold"), fg="#d35400")
    lbl_title.pack(pady=10)

    frame_info = tk.Frame(root)
    frame_info.pack(pady=5, padx=10, fill="x")

    tk.Label(frame_info, text="Mã Bài Hát:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    entry_mabh = tk.Entry(frame_info, width=15)
    entry_mabh.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    tk.Label(frame_info, text="Loại:").grid(row=0, column=2, padx=5, pady=5, sticky="w")

    cbb_loai = ttk.Combobox(frame_info, values=["Rap", "VPop", "Ballad", "Indie", "Bolerro", "Nhạc Vàng", "Nhạc Đỏ", "Du Ca", "KPop", "Jazz"], width=18, state="readonly")
    cbb_loai.grid(row=0, column=3, padx=5, pady=5, sticky="w")

    tk.Label(frame_info, text="Tên Ca Sĩ:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
    entry_casi = tk.Entry(frame_info, width=25)
    entry_casi.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    tk.Label(frame_info, text="Tên Bài Hát:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
    entry_tenbh = tk.Entry(frame_info, width=15)
    entry_tenbh.grid(row=1, column=3, padx=5, pady=5, sticky="w")

    tk.Label(frame_info, text="Giới tính:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
    gender_var = tk.StringVar(value="Nam")
    frame_gender = tk.Frame(frame_info)
    frame_gender.grid(row=2, column=1, sticky="w")
    tk.Radiobutton(frame_gender, text="Nam", variable=gender_var, value="Nam").pack(side=tk.LEFT)
    tk.Radiobutton(frame_gender, text="Nữ", variable=gender_var, value="Nữ").pack(side=tk.LEFT, padx=10)

    tk.Label(frame_info, text="Ngày Phát Hành:").grid(row=2, column=2, padx=5, pady=5, sticky="w")

    date_entry = DateEntry(frame_info, width=16, background="darkblue", 
                       foreground="white", date_pattern="yyyy-mm-dd",
                       year=2000, month=1, day=1)
    date_entry.grid(row=2, column=3, padx=5, pady=5, sticky="w")

    frame_btn = tk.Frame(root)
    frame_btn.pack(pady=15)

    btn_width = 9
    tk.Button(frame_btn, text="Thêm", width=btn_width, command=ThemBaiHat, bg="#2ecc71", fg="white").grid(row=0, column=0, padx=5)
    tk.Button(frame_btn, text="Lưu", width=btn_width, command=LuuBaiHat, bg="#3498db", fg="white").grid(row=0, column=1, padx=5)
    tk.Button(frame_btn, text="Sửa", width=btn_width, command=SuaBaiHat, bg="#f1c40f").grid(row=0, column=2, padx=5)
    tk.Button(frame_btn, text="Hủy", width=btn_width, command=clear_input).grid(row=0, column=3, padx=5)
    tk.Button(frame_btn, text="Xóa", width=btn_width, command=XoaBaiHat, bg="#e74c3c", fg="white").grid(row=0, column=4, padx=5)
    tk.Button(frame_btn, text="Tìm Kiếm", width=btn_width, command=TimKiem, bg="#9b59b6", fg="white").grid(row=0, column=6, padx=5)
    tk.Button(frame_btn, text="Xem Tất Cả", width=btn_width, command=load_data, bg="#95a5a6", fg="white").grid(row=0, column=7, padx=5)
    tk.Button(frame_btn, text="Xuất Excel", width=btn_width, command=XuatExcel, bg="#1abc9c", fg="white").grid(row=0, column=8, padx=5)
    tk.Button(frame_btn, text="Thoát", width=btn_width, command=root.quit).grid(row=0, column=9, padx=5)

    tk.Label(root, text="Danh sách Bài Hát", font=("Arial", 10, "bold")).pack(pady=5, anchor="w", padx=20)

    columns = ("ma_bh", "tenbaihat", "casi", "gioitinh", "ngayphathanh", "loai")
    tree = ttk.Treeview(root, columns=columns, show="headings", height=10)

    tree.heading("ma_bh", text="Mã BH")
    tree.heading("tenbaihat", text="Tên Bài Hát")
    tree.heading("casi", text="Ca Sĩ")
    tree.heading("gioitinh", text="Giới tính")
    tree.heading("ngayphathanh", text="Ngày Phát Hành")
    tree.heading("loai", text="Loại")

    tree.column("ma_bh", width=80, anchor="center")
    tree.column("tenbaihat", width=150)
    tree.column("casi", width=80)
    tree.column("gioitinh", width=60, anchor="center")
    tree.column("ngayphathanh", width=100, anchor="center")
    tree.column("loai", width=120)

    tree.pack(padx=20, pady=5, fill="both", expand=True)

    tree.bind("<Double-1>", SuaBaiHat)


    menubar = tk.Menu(root)
    root.config(menu=menubar)
    file_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="File", menu=file_menu)
    file_menu.add_command(label="Thêm", command=ThemBaiHat)
    file_menu.add_command(label="Lưu", command=LuuBaiHat)
    file_menu.add_command(label="Sửa", command=SuaBaiHat)   
    file_menu.add_command(label="Hủy", command=clear_input)
    file_menu.add_command(label="Xóa", command=XoaBaiHat)
    file_menu.add_command(label="Tìm Kiếm", command=TimKiem)
    file_menu.add_command(label="Xem Tất Cả Bài Hát", command=load_data)
    file_menu.add_command(label="Xuất Excel", command=XuatExcel)
    file_menu.add_command(label="Thoát", command=root.quit)

    load_data()
    clear_input()
    root.mainloop()
